import pandas as pd
import numpy as np
import sqlite3
import statsmodels.api as sm
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
import warnings
warnings.filterwarnings('ignore')

# ── 1. Load data and re-run models ────────────────────────────────────────────
conn = sqlite3.connect('findex_vietnam.db')
df = pd.read_sql('SELECT * FROM findex', conn)
conn.close()
df = df.dropna(subset=['educ', 'anydigpayment'])
y = df['anydigpayment'].astype(int)
X = pd.DataFrame({
    'age': df['age'], 'age_sq': df['age']**2,
    'female': (df['female']==1).astype(int),
    'educ': df['educ'], 'inc_q': df['inc_q'],
    'rural': (df['urbanicity']==1).astype(int),
    'employed': df['emp_in'], 'internet': df['internet_use'],
    'saved': df['saved'], 'borrowed': df['borrowed'],
    'pay_util': df['pay_utilities'],
})
result = sm.Logit(y, sm.add_constant(X.astype(float))).fit(disp=0)
marginal = result.get_margeff()

# ── 2. Build dataframes for each sheet ───────────────────────────────────────

# Sheet 1: Adoption rates
adoption = pd.DataFrame([
    ['Overall account ownership',    68.5],
    ['Digital account ownership',    60.9],
    ['Any digital payment',          60.9],
    ['Mobile account',               36.9],
    ['Merchant digital payment',     49.1],
], columns=['Indicator', 'Rate (%)'])

by_income = pd.DataFrame({
    'Income quintile': ['Q1 (lowest)', 'Q2', 'Q3', 'Q4', 'Q5 (highest)'],
    'n': [203, 189, 198, 207, 203],
    'Digital payment adoption (%)': [23.2, 52.4, 71.7, 75.4, 81.3]
})

by_age = pd.DataFrame({
    'Age band': ['15–24', '25–34', '35–44', '45–54', '55+'],
    'n': [215, 210, 246, 154, 175],
    'Digital payment adoption (%)': [62.8, 83.3, 71.5, 59.1, 18.3]
})

by_educ = pd.DataFrame({
    'Education level': ['Primary', 'Secondary', 'Tertiary'],
    'n': [218, 672, 108],
    'Digital payment adoption (%)': [27.1, 66.7, 93.5]
})

by_urb = pd.DataFrame({
    'Urbanicity': ['Urban', 'Rural'],
    'n': [680, 320],
    'Digital payment adoption (%)': [57.5, 68.1]
})

# Sheet 2: Regression results
var_labels = {
    'const': 'Constant', 'age': 'Age', 'age_sq': 'Age squared',
    'female': 'Female', 'educ': 'Education level', 'inc_q': 'Income quintile',
    'rural': 'Rural location', 'employed': 'Employed',
    'internet': 'Internet user', 'saved': 'Formal saver',
    'borrowed': 'Borrowed (formal)', 'pay_util': 'Pays utilities digitally'
}
reg_df = pd.DataFrame({
    'Variable': [var_labels.get(i, i) for i in result.params.index],
    'Coefficient': result.params.values.round(4),
    'Odds ratio': np.exp(result.params.values).round(4),
    'p-value': result.pvalues.values.round(4),
    'Significant (p<0.05)': ['Yes' if p < 0.05 else 'No'
                              for p in result.pvalues.values]
})

me_labels = {
    'internet': 'Internet user', 'educ': 'Education level',
    'saved': 'Formal saver', 'inc_q': 'Income quintile',
    'age': 'Age', 'age_sq': 'Age squared',
    'borrowed': 'Borrowed (formal)', 'rural': 'Rural location',
    'female': 'Female', 'pay_util': 'Pays utilities digitally',
    'employed': 'Employed'
}
me_df = pd.DataFrame({
    'Variable': [me_labels.get(i, i) for i in X.columns],
    'Marginal effect (pp)': (marginal.margeff * 100).round(2),
    'p-value': marginal.pvalues.round(4),
    'Significant': ['Yes' if p < 0.05 else 'No' for p in marginal.pvalues]
}).sort_values('Marginal effect (pp)', ascending=False).reset_index(drop=True)

# Sheet 3: Segment profiles
gap = df[(df['account']==1) & (df['anydigpayment']==0)]
adopters = df[(df['account']==1) & (df['anydigpayment']==1)]
unbanked = df[df['account']==0]

seg_df = pd.DataFrame({
    'Characteristic': [
        'Sample size (n)', 'Share of total sample (%)',
        'Average age', '% rural', '% female',
        'Average income quintile', 'Average education level',
        '% internet users', '% formal savers'
    ],
    'Next-wave adopters\n(banked, non-digital)': [
        len(gap), round(len(gap)/len(df)*100,1),
        round(gap['age'].mean(),1),
        round((gap['urbanicity']==1).mean()*100,1),
        round((gap['female']==1).mean()*100,1),
        round(gap['inc_q'].mean(),2),
        round(gap['educ'].mean(),2),
        round(gap['internet_use'].mean()*100,1),
        round(gap['saved'].mean()*100,1),
    ],
    'Full digital adopters\n(banked + digital)': [
        len(adopters), round(len(adopters)/len(df)*100,1),
        round(adopters['age'].mean(),1),
        round((adopters['urbanicity']==1).mean()*100,1),
        round((adopters['female']==1).mean()*100,1),
        round(adopters['inc_q'].mean(),2),
        round(adopters['educ'].mean(),2),
        round(adopters['internet_use'].mean()*100,1),
        round(adopters['saved'].mean()*100,1),
    ],
    'Unbanked': [
        len(unbanked), round(len(unbanked)/len(df)*100,1),
        round(unbanked['age'].mean(),1),
        round((unbanked['urbanicity']==1).mean()*100,1),
        round((unbanked['female']==1).mean()*100,1),
        round(unbanked['inc_q'].mean(),2),
        round(unbanked['educ'].mean(),2),
        round(unbanked['internet_use'].mean()*100,1),
        round(unbanked['saved'].mean()*100,1),
    ],
})

# ── 3. Write to Excel ─────────────────────────────────────────────────────────
output_file = 'Techcombank_Digital_Adoption_Analysis.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    adoption.to_excel(writer, sheet_name='1. Adoption Rates',
                      startrow=7, index=False)
    by_income.to_excel(writer, sheet_name='1. Adoption Rates',
                       startrow=15, index=False)
    by_age.to_excel(writer, sheet_name='1. Adoption Rates',
                    startrow=23, index=False)
    by_educ.to_excel(writer, sheet_name='1. Adoption Rates',
                     startrow=31, index=False)
    by_urb.to_excel(writer, sheet_name='1. Adoption Rates',
                    startrow=37, index=False)

    reg_df.to_excel(writer, sheet_name='2. Regression Results',
                    startrow=7, index=False)
    me_df.to_excel(writer, sheet_name='2. Regression Results',
                   startrow=7+len(reg_df)+4, index=False)

    seg_df.to_excel(writer, sheet_name='3. Segment Profiles',
                    startrow=7, index=False)

# ── 4. Style the workbook ─────────────────────────────────────────────────────
wb = load_workbook(output_file)

# Colour palette
TCB_RED    = 'C8102E'   # Techcombank red
TCB_DARK   = '1A1A2E'   # dark navy for headers
LIGHT_GRAY = 'F5F5F5'
MID_GRAY   = 'E0E0E0'
WHITE      = 'FFFFFF'
GREEN_LIGHT= 'E8F5E9'
RED_LIGHT  = 'FFEBEE'

header_font  = Font(name='Calibri', bold=True, color=WHITE, size=11)
title_font   = Font(name='Calibri', bold=True, color=TCB_DARK, size=13)
section_font = Font(name='Calibri', bold=True, color=TCB_RED, size=11)
body_font    = Font(name='Calibri', size=10)
header_fill  = PatternFill('solid', fgColor=TCB_DARK)
alt_fill     = PatternFill('solid', fgColor=LIGHT_GRAY)
center       = Alignment(horizontal='center', vertical='center', wrap_text=True)
left         = Alignment(horizontal='left',   vertical='center', wrap_text=True)
thin         = Side(style='thin', color='D0D0D0')
thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_sheet(ws, tables):
    """Apply consistent styling to a worksheet."""
    ws.sheet_view.showGridLines = False

    # Title block
    ws['A1'] = 'Techcombank Digital Payment Adoption Analysis'
    ws['A1'].font = Font(name='Calibri', bold=True, color=TCB_RED, size=15)
    ws['A2'] = 'Vietnam | World Bank Global Findex 2021 | n = 998'
    ws['A2'].font = Font(name='Calibri', color='666666', size=10)
    ws['A3'] = 'Prepared for: Consumer Insights Report | May 2026'
    ws['A3'].font = Font(name='Calibri', color='666666', size=10)

    for (start_row, df_ref, title) in tables:
        # Section title
        title_cell = ws.cell(row=start_row-1, column=1, value=title)
        title_cell.font = section_font

        # Header row
        n_cols = len(df_ref.columns)
        for col_idx, col_name in enumerate(df_ref.columns, start=1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        # Data rows
        for row_idx in range(start_row+1, start_row+len(df_ref)+1):
            fill = alt_fill if (row_idx - start_row) % 2 == 0 else PatternFill('solid', fgColor=WHITE)
            for col_idx in range(1, n_cols+1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = body_font
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = center

    # Column widths
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

# Sheet 1
ws1 = wb['1. Adoption Rates']
style_sheet(ws1, [
    (8,  adoption,  'Overview: key adoption indicators'),
    (16, by_income, 'Digital payment adoption by income quintile'),
    (24, by_age,    'Digital payment adoption by age band'),
    (32, by_educ,   'Digital payment adoption by education level'),
    (38, by_urb,    'Digital payment adoption by urbanicity'),
])

# Highlight significant cells in Sheet 1
for row in ws1.iter_rows(min_row=9, max_row=13, min_col=2, max_col=2):
    for cell in row:
        try:
            val = float(cell.value)
            if val >= 70:
                cell.fill = PatternFill('solid', fgColor=GREEN_LIGHT)
            elif val <= 30:
                cell.fill = PatternFill('solid', fgColor=RED_LIGHT)
        except Exception:
            pass

# Sheet 2
ws2 = wb['2. Regression Results']
style_sheet(ws2, [
    (8,  reg_df, 'Logistic regression coefficients (outcome: any digital payment)'),
    (8+len(reg_df)+4+1, me_df, 'Average marginal effects (percentage points)'),
])

# Colour-code significance in regression sheet
sig_col = reg_df.columns.get_loc('Significant (p<0.05)') + 1
for row in ws2.iter_rows(min_row=9, max_row=8+len(reg_df), min_col=sig_col, max_col=sig_col):
    for cell in row:
        if cell.value == 'Yes':
            cell.fill = PatternFill('solid', fgColor=GREEN_LIGHT)
            cell.font = Font(name='Calibri', size=10, bold=True, color='1B5E20')
        elif cell.value == 'No':
            cell.fill = PatternFill('solid', fgColor=RED_LIGHT)
            cell.font = Font(name='Calibri', size=10, color='B71C1C')

# Sheet 3
ws3 = wb['3. Segment Profiles']
style_sheet(ws3, [
    (8, seg_df, 'Consumer segment profiles: next-wave adopters vs. full adopters vs. unbanked'),
])

wb.save(output_file)
print(f"Workbook saved: {output_file}")
print("Sheets: 1. Adoption Rates | 2. Regression Results | 3. Segment Profiles")
